import customtkinter
import time
import openpyxl
from pathlib import Path

from PneumoCVUTFBMI.DeviceLoader import DeviceLoader

workbook = openpyxl.Workbook()
worksheet = workbook.active
Minuly_Sval = 0
Zvoleny_Sval = 0
y = 0
x = 1
pokracovat = True

dl = DeviceLoader()

b1 = dl.getBoard1()
b2 = dl.getBoard2()
b3 = dl.getBoard3()
b4 = dl.getBoard4()
b5 = dl.getBoard5()

# Modes: "System" (standard), "Dark", "Light"
customtkinter.set_appearance_mode("System")
# Themes: "blue" (standard), "green", "dark-blue"
customtkinter.set_default_color_theme("blue")

class adminWindow(customtkinter.CTkToplevel):
    def __init__(self, mainwindow, *args, **kwargs):
        super().__init__(*args, **kwargs)

        # self.technicka_nula() # tohle se bude řešit teprve až to bude
        # Set window size
        width = 300
        height = 150
        
        # Automatic calculation position of window based on monitor size
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        screen_x = (screen_width - width) // 2
        screen_y = (screen_height - height) // 2
        self.title("admin")

        # Set size a place parametrs
        self.geometry(f"{width}x{height}+{screen_x}+{screen_y}")

        self.measure_button = customtkinter.CTkButton(self, text="Měření", command= self.mesureWindow) # button for login
        self.measure_button.pack(side="top", pady = 10)

        self.db_button = customtkinter.CTkButton(self, text="Databáze", command= self.databaseWindow) # button for login
        self.db_button.pack(side="top", pady = 10)

        self.home_button = customtkinter.CTkButton(self, text="Hlavní stránka", command= self.back2Main) # button for login
        self.home_button.pack(side="bottom", pady = 10)

        self.mainwindow = mainwindow
    
    def mesureWindow(self):
        self.toplevel = mesuremenWindow(self,self)
        self.withdraw()

    def databaseWindow(self):
        self.toplevel = databaseWindow(self,self)
        self.withdraw()

    def back2Main(self):
        self.withdraw()
        self.mainwindow.deiconify()

class mesuremenWindow(customtkinter.CTkToplevel):
    def __init__(self, mainwindow, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Set window size
        width = 540
        height = 220
        
        # Automatic calculation position of window based on monitor size
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        screen_x = (screen_width - width) // 2
        screen_y = (screen_height - height) // 2
        self.title("mesure")

        # Set size a place parametrs
        self.geometry(f"{width}x{height}+{screen_x}+{screen_y}")

        # create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(
            self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=6, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        self.logo_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="Výběr svalu", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        # create radiobutton frame
        self.radio_var = customtkinter.IntVar(value=0)

        self.radio_button_1 = customtkinter.CTkRadioButton(
            master=self.sidebar_frame, variable=self.radio_var, value=1, command=self.radiobutton_event, text="Sval 1")
        self.radio_button_1.grid(row=1, column=0)
        self.radio_button_2 = customtkinter.CTkRadioButton(
            master=self.sidebar_frame, variable=self.radio_var, value=2, command=self.radiobutton_event, text="Sval 2")
        self.radio_button_2.grid(row=2, column=0)
        self.radio_button_3 = customtkinter.CTkRadioButton(
            master=self.sidebar_frame, variable=self.radio_var, value=3, command=self.radiobutton_event, text="Sval 3")
        self.radio_button_3.grid(row=3, column=0)
        self.radio_button_4 = customtkinter.CTkRadioButton(
            master=self.sidebar_frame, variable=self.radio_var, value=4, command=self.radiobutton_event, text="Sval 4")
        self.radio_button_4.grid(row=4, column=0)
        self.radio_button_5 = customtkinter.CTkRadioButton(
            master=self.sidebar_frame, variable=self.radio_var, value=5, command=self.radiobutton_event, text="Sval 5")
        self.radio_button_5.grid(row=5, column=0)

        # Apearence mode options
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Light", "Dark", "System"],
                                                                       command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(
            row=6, column=0, padx=20, pady=(10, 10))
        self.appearance_mode_optionemenu.set("System")

        # create main entry and button
        self.entrySpeed = customtkinter.CTkEntry(
            self, placeholder_text="Speed")
        self.entrySpeed.grid(row=2, column=1,  padx=(
            20, 0), pady=(20, 20), sticky="nsew")

        self.entrySteps = customtkinter.CTkEntry(
            self, placeholder_text="Steps")
        self.entrySteps.grid(row=2, column=2, padx=(
            20, 0), pady=(20, 20), sticky="nsew")

        self.entry = customtkinter.CTkEntry(self, placeholder_text="Hw Value")
        self.entry.grid(row=3, column=1, padx=(20, 0),
                        pady=(20, 20), sticky="nsew")

        self.main_button_1 = customtkinter.CTkButton(
            self, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.hw_value, text="krok")
        self.main_button_1.grid(row=3, column=2, padx=(
            20, 20), pady=(20, 20), sticky="nsew")
        # tohle nevím proč nefunguje je to funkce na přidávání pomocí enteru
        self.bind("<Return>", lambda event: self.hw_value())
        self.bind("<KP_Enter>", lambda event: self.hw_value())

        self.main_button_2 = customtkinter.CTkButton(master=self, fg_color="transparent", border_width=2, text_color=(
            "gray10", "#DCE4EE"), command=self.change_pokracovat, text ="pokračovat")
        self.main_button_2.grid(row=4, column=2, padx=(
            20, 20), pady=(20, 20), sticky="nsew")
        self.main_button_2.configure(state='disabled')


        self.main_button_3 = customtkinter.CTkButton(
            self, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.back2AdminWindow, text="Hl admin")
        self.main_button_3.grid(row=4, column=1, padx=(
            20, 20), pady=(20, 20), sticky="nsew")

        self.mainwindow = mainwindow

    def back2AdminWindow(self):
        desktop_path = Path.home() / "Desktop"
        workbook.save(desktop_path / f"sval{Minuly_Sval}.xlsx")
        global x
        
        if 'b_objects' in globals() and b_objects is not None:
            if Minuly_Sval in b_objects:
                for i in range(1, x):
                    b_objects[Minuly_Sval].go_backward(Speed, Steps)
                    x=x-1
                    time.sleep(3)
        


        self.withdraw()
        self.mainwindow.deiconify()


    def change_pokracovat(self):
        global pokracovat
        pokracovat = True
        self.main_button_2.configure(state='disabled')

    def hw_value(self):
        global x
        global y
        global Speed
        global Steps
        global Minuly_Sval
        global pokracovat
        
        

        # Vytvoření slovníku pro objekty b1, b2, atd.
        global b_objects
        b_objects = {
            1: b1,
            2: b2,
            3: b3,
            4: b4,
            5: b5
        }

        if Minuly_Sval != Zvoleny_Sval:
            if Minuly_Sval in b_objects:
                for i in range(1, x):
                    b_objects[Minuly_Sval].go_backward(Speed, Steps)
                    time.sleep(3)

            worksheet.cell(row=1, column=1, value="Počet krokůmotoru")
            worksheet.cell(row=1, column=2, value="Hodnota v mV")
            worksheet.cell(row=1, column=3, value="Hodnota fyzického senzoru")
            worksheet.cell(row=1, column=4, value="Rychlost")
            desktop_path = Path.home() / "Desktop"
            workbook.save(desktop_path / f"sval{Minuly_Sval}.xlsx")

            print("změna svalu")
            x = 1
            y = 0
            Minuly_Sval = Zvoleny_Sval
            pokracovat = False
            self.main_button_2.configure(state='enable')

        if pokracovat == True:
            if Zvoleny_Sval in b_objects:
                b_object = b_objects[Zvoleny_Sval]

                b_object.on()
                Speed = self.entrySpeed.get()
                Steps = self.entrySteps.get()
                # Tady se nám počítá kolik udělal motor celkově kroků
                y = y + int(Steps)
                x = x + 1
                b_object.go_forward(Steps, Speed)
                time.sleep(3)
                worksheet.cell(row=x, column=1, value=y)  # počet kroků motoru
                # Zde se načítá ze senzoru může být problém
                worksheet.cell(row=x, column=2, value=str(b_object.readA0()))
                # hodnota v mv
                worksheet.cell(row=x, column=3, value=self.entry.get())
                # hodnota Rychlosti
                worksheet.cell(row=x, column=4, value=Speed)
                Minuly_Sval = Zvoleny_Sval
                self.entry.delete(0, 10000)
                print(x)
            else:
                print("Neplatný Zvoleny_Sval")

    def open_input_dialog_event(self):
        dialog = customtkinter.CTkInputDialog(
            text="Type in a number:", title="CTkInputDialog")
        print("CTkInputDialog:", dialog.get_input())

    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def getRadiVar(self):
        print("Počáteční hodnota:", self.radio_var.get())

    def radiobutton_event(self):
        global Zvoleny_Sval

        selected_value = self.radio_var.get()

        if 1 <= selected_value <= 5:
            print(f"Sval {selected_value}")
            Zvoleny_Sval = selected_value
        else:
            print("Neznámý sval")

    def technicka_nula(self):
        global svaly
        svaly = {
                0: b1,
                1: b2,
                2: b3,
                3: b4,
                4: b5
            }

        brokenMuscle = 3

        while True:  
            results = [sval.readA0() for sval in svaly.values()]

            all_in_range = all(595 <= result <= 610 for i, result in enumerate(results) if i != brokenMuscle)

            if all_in_range:
                break  
            
            for i in range(5):
                if i==brokenMuscle:
                    continue
                
                if results[i] >= 615:
                    svaly[i].go_backward(10, 10)
                elif results[i] <= 595:
                    svaly[i].go_forward(10, 10)


            time.sleep(3)

class databaseWindow(customtkinter.CTkToplevel):
    def __init__(self, mainwindow, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Set window size
        width = 300
        height = 150
        
        # Automatic calculation position of window based on monitor size
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        screen_x = (screen_width - width) // 2
        screen_y = (screen_height - height) // 2
        self.title("database")

        # Set size a place parametrs
        self.geometry(f"{width}x{height}+{screen_x}+{screen_y}")

        self.mainwindow = mainwindow

        self.back2Admin_button = customtkinter.CTkButton(self, text="Hl Admin", command= self.back2AdminWindow) # button for login
        self.back2Admin_button.pack(side="top", pady = 10)

    def back2AdminWindow(self):
        self.withdraw()
        self.mainwindow.deiconify()